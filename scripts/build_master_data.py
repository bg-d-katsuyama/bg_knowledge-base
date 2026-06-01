"""抽出済み知見 JSON から土壌 R&D マスターデータ（xlsx + CSV）を生成する.

Anthropic API は使わない。知見抽出は Claude Code（人手）が行い、結果を
JSON で受け取って、承認済みフォーマット（master_data_sample 準拠）の
スプレッドシート用ファイルに整形するだけの決定論的スクリプト。

入力 JSON 形式（配列。各要素のキーは任意・欠損は空文字で補完）:
    [
      {
        "title": "...", "content": "...", "domain": "堆肥",
        "tags": "牛糞, 含水率", "speaker": "", "affiliation": "BG",
        "confidence": "高", "verification_status": "実証済み",
        "source_occurred": "2026-05-28",
        "source_title": "腐植性堆肥製造マニュアル_v1.3",
        "source_url": "https://www.notion.so/...",
        "source_kb_id": "312be8e0..."
      }, ...
    ]

実行方法:
    uv run python scripts/build_master_data.py logs/insights.json \\
        [--out-dir output] [--basename soil_rd_master]
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# 承認済みフォーマット（docs/master_data_sample.md 準拠）の列
COLUMNS: list[tuple[str, str]] = [
    ("知見ID", "_row_id"),
    ("知見タイトル", "title"),
    ("内容", "content"),
    ("領域分類", "domain"),
    ("関連タグ", "tags"),
    ("発言者", "speaker"),
    ("所属", "affiliation"),
    ("信頼度", "confidence"),
    ("検証ステータス", "verification_status"),
    ("発生日", "source_occurred"),
    ("出典タイトル", "source_title"),
    ("出典URL", "source_url"),
    ("出典KB_ID", "source_kb_id"),
]


def _normalize_tags(value: Any) -> str:
    """tags がリストでもカンマ区切り文字列でも、カンマ区切り文字列に揃える."""
    if isinstance(value, list):
        return ", ".join(str(v).strip() for v in value if str(v).strip())
    return str(value or "")


def _row_values(insight: dict[str, Any], row_id: str) -> list[str]:
    out: list[str] = []
    for _header, key in COLUMNS:
        if key == "_row_id":
            out.append(row_id)
        elif key == "tags":
            out.append(_normalize_tags(insight.get("tags")))
        else:
            out.append(str(insight.get(key, "") or ""))
    return out


def _write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "土壌R&Dマスター"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row)

    # 内容列だけ折り返し・広め、他はそこそこの幅に
    widths = {"知見ID": 10, "知見タイトル": 30, "内容": 60, "領域分類": 10}
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(header, 16)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("insights_json")
    parser.add_argument("--out-dir", default="output")
    parser.add_argument("--basename", default="soil_rd_master")
    args = parser.parse_args()

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    data = json.loads(Path(args.insights_json).read_text(encoding="utf-8"))
    if not isinstance(data, list):
        print("入力 JSON はトップレベルが配列である必要があります", file=sys.stderr)
        return 1

    headers = [h for h, _ in COLUMNS]
    rows = [_row_values(ins, f"K-{i:05d}") for i, ins in enumerate(data, 1)]

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = out_dir / f"{args.basename}_{ts}.csv"
    xlsx_path = out_dir / f"{args.basename}_{ts}.xlsx"

    _write_csv(csv_path, headers, rows)
    _write_xlsx(xlsx_path, headers, rows)

    print(f"知見件数: {len(rows)}")
    print(f"CSV : {csv_path}")
    print(f"xlsx: {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
